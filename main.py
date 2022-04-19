from win32com.client import Dispatch
import os
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from operator import add

# Столбцы, которые переносятся в итоговый файл без изменения
column_indexes = (1, 2, 3, 4, 5, 6, 7, 8, 9, 12, 16, 17, 19, 20, 22, 23, 25, 26)

# Данные из ОТК, которые необходимо сопоставить с итоговым файлом
otk_book = openpyxl.load_workbook('data/otk.xlsx', data_only=True)
otk_sheet = otk_book.active
otk_dict = {}  # Словарь с ключами номенклатур и значениями кол-ва не оприходованных шт.
for r in range(1, 250):
    otk_dict[otk_sheet.cell(row=r, column=1).value] = otk_sheet.cell(row=r, column=2).value

file_folder_path_dict = {
    # 'abutments': ('abutments_info.xlsx', 'abutments_unsh.xlsx', 'Абатменты'),
    # 'abutments_fired': ('abutments_fired_info.xlsx', 'abutments_fired_unsh.xlsx', 'Абатменты_выжигаемые'),
    # 'analog': ('analog_info.xlsx', 'analog_unsh.xlsx', 'Аналоги'),
    # 'blanks': ('blanks_info.xlsx', 'blanks_unsh.xlsx', 'Заготовки'),
    # 'formers': ('formers_info.xlsx', 'formers_unsh.xlsx', 'Формирователи'),
    # 'implants': ('implants_info.xlsx', 'implants_stock.xlsx', 'Импланты'),
    # 'scan_body': ('scan_body_info.xlsx', 'scan_body_unsh.xlsx', 'Скан_боди'),
    # 'screws': ('screws_info.xlsx', 'screws_unsh.xlsx', 'Винты'),
    # 'sleeve': ('sleeve_info.xlsx', 'sleeve_unsh.xlsx', 'Втулка'),
    'titanium_base': ('titanium_base_info.xlsx', 'titanium_base_unsh.xlsx', 'Титановые_основы'),
    'transfers': ('transfers_info.xlsx', 'transfers_unsh.xlsx', 'Трансферы'),
}


def create_date_dict():
    month_dict = {
        'марта': '03',
        'апр.': '04',
        'мая': '05'
    }

    global date_dict
    date_dict = {}
    for row in row_indexes:
        if input_sheet.cell(row=row, column=27).value is None:
            date_dict[row] = None
        else:
            dates = []
            flag = True
            for col in range(29, input_sheet.max_column+1):
                if flag:
                    if input_sheet.cell(row=row, column=col).value is None:
                        continue
                    else:
                        dates.append([str(input_sheet.cell(row=1, column=col).value)])
                        flag = False
                else:
                    if input_sheet.cell(row=row, column=col).value is None:
                        flag = True
                        continue
                    else:
                        dates[-1].append(str(input_sheet.cell(row=1, column=col).value))

            result = ''
            for item in dates:
                if result:
                    if len(item) == 1:
                        result += f', {item[0].split()[0]}.{month_dict.get(item[0].split()[1])}'
                    else:
                        if item[0].split()[-1] == item[-1].split()[-1]:
                            result += f', {item[0].split()[0]}-' \
                                      f'{item[-1].split()[0]}.' \
                                      f'{month_dict.get(item[-1].split()[1])}'
                        else:
                            result += f', {item[0].split()[0]}.{month_dict.get(item[0].split()[1])}-' \
                                      f'{item[-1].split()[0]}.{month_dict.get(item[-1].split()[1])}'
                else:
                    if len(item) == 1:
                        result += f'{item[0].split()[0]}.{month_dict.get(item[0].split()[1])}'
                    else:
                        if item[0].split()[-1] == item[-1].split()[1]:
                            result += f'{item[0].split()[0]}-' \
                                      f'{item[-1].split()[0]}.' \
                                      f'{month_dict.get(item[-1].split()[1])}'
                        else:
                            result += f'{item[0].split()[0]}.{month_dict.get(item[0].split()[1])}-' \
                                      f'{item[-1].split()[0]}.{month_dict.get(item[-1].split()[1])}'

            date_dict[row] = result


def read_input_files(main_file, unshipped_file):
    input_book = openpyxl.load_workbook(main_file)
    global input_sheet
    input_sheet = input_book.active

    # Отбор не архивных номенклатур в main_file
    global row_indexes
    row_indexes = []
    for row in range(4, input_sheet.max_row + 1):
        if input_sheet.cell(row=row, column=6).value != 'Да' and input_sheet.cell(row=row, column=7).value != 'Да':
            row_indexes.append(row)

    # Отбор арх. поз. titanium_base, которые надо учесть
    if main_file == 'data/titanium_base/titanium_base_info.xlsx':
        tb_actual_book = openpyxl.load_workbook('data/titanium_base/titanium_base_actual.xlsx')
        tb_actual_sheet = tb_actual_book.active
        tb_actual_dict = {}
        for r in range(2, 20):
            tb_actual_dict[tb_actual_sheet.cell(row=r, column=1).value] = \
                tb_actual_dict.get(tb_actual_sheet.cell(row=r, column=1).value, []) + \
                [tb_actual_sheet.cell(row=r, column=2).value]

        global tb_ar_dict
        tb_ar_dict = {}
        for row in range(4, input_sheet.max_row + 1):
            if input_sheet.cell(row=row, column=5).value in list(tb_actual_dict.values())[0]:
                info = []
                for col in (9, 12, 16, 17, 19, 20, 22, 23, 25, 26):
                    if input_sheet.cell(row=row, column=col).value is None:
                        x = 0
                    else:
                        x = int(input_sheet.cell(row=row, column=col).value)
                    info.append(x)

                tb_ar_dict[list(tb_actual_dict.keys())[0]] = \
                    list(
                        map(
                            add,
                            tb_ar_dict.get(list(tb_actual_dict.keys())[0], [0 for _ in range(10)]),
                            info
                        )
                    )
            elif input_sheet.cell(row=row, column=5).value in list(tb_actual_dict.values())[1]:
                info = []
                for col in (9, 12, 16, 17, 19, 20, 22, 23, 25, 26):
                    if input_sheet.cell(row=row, column=col).value is None:
                        x = 0
                    else:
                        x = int(input_sheet.cell(row=row, column=col).value)
                    info.append(x)

                tb_ar_dict[list(tb_actual_dict.keys())[1]] = \
                    list(
                        map(
                            add,
                            tb_ar_dict.get(list(tb_actual_dict.keys())[1], [0 for _ in range(10)]),
                            info
                        )
                    )
            elif input_sheet.cell(row=row, column=5).value in list(tb_actual_dict.values())[2]:
                info = []
                for col in (9, 12, 16, 17, 19, 20, 22, 23, 25, 26):
                    if input_sheet.cell(row=row, column=col).value is None:
                        x = 0
                    else:
                        x = int(input_sheet.cell(row=row, column=col).value)
                    info.append(x)

                tb_ar_dict[list(tb_actual_dict.keys())[2]] = \
                    list(
                        map(
                            add,
                            tb_ar_dict.get(list(tb_actual_dict.keys())[2], [0 for _ in range(10)]),
                            info
                        )
                    )
            elif input_sheet.cell(row=row, column=5).value in list(tb_actual_dict.values())[3]:
                info = []
                for col in (9, 12, 16, 17, 19, 20, 22, 23, 25, 26):
                    if input_sheet.cell(row=row, column=col).value is None:
                        x = 0
                    else:
                        x = int(input_sheet.cell(row=row, column=col).value)
                    info.append(x)

                tb_ar_dict[list(tb_actual_dict.keys())[3]] = \
                    list(
                        map(
                            add,
                            tb_ar_dict.get(list(tb_actual_dict.keys())[3], [0 for _ in range(10)]),
                            info
                        )
                    )

    # Дополнительный исключения номенклатур
    row_indexes_copy = row_indexes.copy()
    if main_file == 'data/screws/screws_info.xlsx':
        for i, row in enumerate(row_indexes_copy):
            if 'блистер' in input_sheet.cell(row=row, column=5).value.lower():
                del row_indexes[row_indexes.index(row)]
            elif 'упак' in input_sheet.cell(row=row, column=5).value.lower():
                del row_indexes[row_indexes.index(row)]
            elif 'проб' in input_sheet.cell(row=row, column=5).value.lower():
                del row_indexes[row_indexes.index(row)]
    elif main_file == 'data/analog/analog_info.xlsx':
        for i, row in enumerate(row_indexes_copy):
            if 'нерж' not in input_sheet.cell(row=row, column=5).value.lower():
                del row_indexes[row_indexes.index(row)]
    elif main_file == 'data/scan_body/scan_body_info.xlsx':
        for i, row in enumerate(row_indexes_copy):
            if 'нерж' in input_sheet.cell(row=row, column=5).value.lower() and \
                    'б' in input_sheet.cell(row=row, column=5).value.split()[0]:
                del row_indexes[row_indexes.index(row)]
            elif 'латунь' in input_sheet.cell(row=row, column=5).value.lower() and \
                    'б' in input_sheet.cell(row=row, column=5).value.split()[0]:
                del row_indexes[row_indexes.index(row)]
    elif main_file == 'data/titanium_base/titanium_base_info.xlsx':
        for i, row in enumerate(row_indexes_copy):
            if 'струк' in input_sheet.cell(row=row, column=5).value.lower():
                del row_indexes[row_indexes.index(row)]
            elif '2к' in input_sheet.cell(row=row, column=5).value:
                del row_indexes[row_indexes.index(row)]
            elif 'кат2' in input_sheet.cell(row=row, column=5).value.split()[0]:
                del row_indexes[row_indexes.index(row)]

    # Создание словаря с неотгруженными номенклатурами из unshipped_file
    global unshipped_dict
    if unshipped_file != 'data/implants/implants_stock.xlsx':
        unshipped_dict = {}
        unshipped_book = openpyxl.load_workbook(unshipped_file)
        unshipped_sheet = unshipped_book.active

        if unshipped_sheet.max_column == 8:
            for row in range(7, unshipped_sheet.max_row + 1):
                unshipped_dict[unshipped_sheet.cell(row=row, column=1).value] = \
                    int(unshipped_sheet.cell(row=row, column=8).value)
            print(f'{unshipped_file} is loaded')
        else:
            print(f'{unshipped_file} have error: max_column != 8:')
    else:
        unshipped_dict = {}

    # Создание словаря для имплантов
    if unshipped_file == 'data/implants/implants_stock.xlsx':
        global stock_dict
        stock_dict = {}
        stock_book = openpyxl.load_workbook(unshipped_file)
        stock_sheet = stock_book.active

        for row in range(1, stock_sheet.max_row + 1):
            stock_dict[stock_sheet.cell(row=row, column=1).value] = \
                stock_sheet.cell(row=row, column=2).value
        print(f'{unshipped_file} is loaded')

    # Создание словаря с датами пр-ва
    create_date_dict()

    print(f'{main_file} is loaded')


def create_sheet_header(sheet_name):
    # Словарь для стилизации шапки
    headers_dict = {
        1: 'Тип',
        2: 'Линейка',
        3: 'Система',
        4: 'Разм',
        5: 'Номенклатура',
        6: 'Арх ном',
        7: 'Арх кат',
        8: 'Карт кат',
        9: 'Остаток',
        10: 'Расход общий',
        19: 'ПЛАН',
        20: 'Произведено / неоприходовано',
        21: 'Непроизведено / в плане',
        22: 'Неотгружено по опт. заявкам'
    }

    sheet_name.cell(row=1, column=5).value = f'Конец периода: {date_stop} 23:59:59'
    sheet_name.cell(row=1, column=5).font = Font(name='Arial', bold=False, size=8)

    sheet_name.cell(row=2, column=5).value = f'Начало периода: {date_start} 00:00:00'
    sheet_name.cell(row=2, column=5).font = Font(name='Arial', bold=False, size=8)

    for key in headers_dict.keys():
        if 9 <= key <= 10:
            sheet_name.merge_cells(start_row=5, start_column=int(key), end_row=6, end_column=int(key))
            sheet_name.cell(row=5, column=int(key)).value = headers_dict[key]
        else:
            sheet_name.merge_cells(start_row=4, start_column=int(key), end_row=6, end_column=int(key))
            sheet_name.cell(row=4, column=int(key)).value = headers_dict[key]

    sheet_name.merge_cells(start_row=4, start_column=9, end_row=4, end_column=10)
    sheet_name.cell(row=4, column=9).value = 'Итого'
    sheet_name.merge_cells(start_row=4, start_column=11, end_row=4, end_column=14)
    sheet_name.cell(row=4, column=11).value = 'ОСНОВНЫЕ СКЛАДЫ'
    sheet_name.merge_cells(start_row=4, start_column=15, end_row=4, end_column=18)
    sheet_name.cell(row=4, column=15).value = 'ПРОЧИЕ СКЛАДЫ'

    for c in [11, 15]:
        sheet_name.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c+1)
        sheet_name.cell(row=5, column=c).value = 'ОСТ'

    for c in [13, 17]:
        sheet_name.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c+1)
        sheet_name.cell(row=5, column=c).value = 'РАСХ'

    for c in range(11, 19):
        if c % 2 != 0:
            sheet_name.cell(row=6, column=c).value = 'ИЗД'
        else:
            sheet_name.cell(row=6, column=c).value = 'К/Т'

    for row in range(4, 7):
        for c in range(1, 23):
            sheet_name.cell(row=row, column=c).style = 'header'

    for c in range(1, 23):
        if c <= 4:
            sheet_name.column_dimensions[get_column_letter(c)].width = 9
        elif c == 5:
            sheet_name.column_dimensions[get_column_letter(c)].width = 90
        elif 6 <= c <= 8:
            sheet_name.column_dimensions[get_column_letter(c)].width = 7.5
        elif 9 <= c <= 18:
            sheet_name.column_dimensions[get_column_letter(c)].width = 8.25
        else:
            sheet_name.column_dimensions[get_column_letter(c)].width = 20

    sheet_name.column_dimensions.group('F', 'H', hidden=True)
    sheet_name.freeze_panes = sheet_name.cell(row=7, column=6)


def fill_small_stock(sheet_name, start_row):
    for row in range(start_row, sheet_name.max_row + 1):
        if sheet_name.cell(row=row, column=9).value is None:
            continue
        elif sheet_name.cell(row=row, column=10).value >= sheet_name.cell(row=row, column=9).value:
            sheet_name.cell(row=row, column=5).fill = PatternFill("solid", fgColor="FFCCCC")
            for c in [9, 10]:
                sheet_name.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFCCCC")
                sheet_name.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
        elif sheet_name.cell(row=row, column=13).value != 0:
            if (sheet_name.cell(row=row, column=11).value / sheet_name.cell(row=row, column=13).value) < 1:
                sheet_name.cell(row=row, column=5).fill = PatternFill("solid", fgColor="CCECFF")
                for c in [11, 13]:
                    sheet_name.cell(row=row, column=c).fill = PatternFill("solid", fgColor="CCECFF")
                    sheet_name.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
            elif 1 <= (sheet_name.cell(row=row, column=11).value / sheet_name.cell(row=row, column=13).value) < 2.5:
                sheet_name.cell(row=row, column=5).fill = PatternFill("solid", fgColor="CCFFCC")
                for c in [11, 13]:
                    sheet_name.cell(row=row, column=c).fill = PatternFill("solid", fgColor="CCFFCC")
                    sheet_name.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')


def create_sheet_result(sheet_name, start_row, end_row):
    rows_dict = {
        9: 'I',
        10: 'J',
        11: 'K',
        12: 'L',
        13: 'M',
        14: 'N',
        15: 'O',
        16: 'P',
        17: 'Q',
        18: 'R',
        20: 'T',
        21: 'U',
        22: 'V',
        23: 'W'
    }

    for col in range(1, sheet_name.max_column+1):
        cell = sheet_name.cell(row=end_row, column=col)
        cell.style = 'header'
        if col == 5:
            cell.value = 'Итого'
        elif 9 <= col <= 18 or 20 <= col:
            cell.value = f'=SUM({rows_dict[col]}{start_row}:{rows_dict[col]}{end_row-1})'
            cell.number_format = '# ##0'


def separation_nomenclatures(sheet_name, start_row):
    # Будем разделять номенклатуры по их линейке, а также объединять пары с одинаковыми размерами в жирные границы
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")

    row_indexes_for_new_line = []
    for row in range(start_row, sheet_name.max_row + 1):
        if sheet_name.cell(row=row, column=3).value == sheet_name.cell(row=(row + 1), column=3).value:
            continue
        else:
            row_indexes_for_new_line.append(row + 1)

    for i in enumerate(row_indexes_for_new_line[:-1:]):
        sheet_name.insert_rows(idx=(i[1] + i[0]))
        for col in range(1, sheet_name.max_column+1):
            cell = sheet_name.cell(row=(i[1] + i[0]), column=col)
            if col != 20 and col != 21:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            else:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")

    border_highlighting_list = []
    for i in enumerate(row_indexes_for_new_line):
        flag = True
        for row in range(start_row, i[1] + i[0] - 1):
            if flag:
                if sheet_name.cell(row=row, column=4).value == sheet_name.cell(row=(row + 1), column=4).value:
                    continue
                else:
                    flag = False
                    stop_row = row
                    border_highlighting_list.append([start_row, stop_row])
                    start_row = row + 1
            else:
                if sheet_name.cell(row=row, column=4).value == sheet_name.cell(row=(row + 1), column=4).value and\
                        sheet_name.cell(row=(row + 2), column=4).value is not None:
                    continue
                elif sheet_name.cell(row=row, column=4).value == sheet_name.cell(row=(row + 1), column=4).value and\
                        sheet_name.cell(row=(row + 2), column=4).value is None:
                    flag = True
                    stop_row = row + 1
                    border_highlighting_list.append([start_row, stop_row])
                    start_row = row + 1
                else:
                    stop_row = row
                    border_highlighting_list.append([start_row, stop_row])
                    start_row = row + 1
        start_row = i[1] + i[0] + 1

    # Закрашиваем границы отобранных ячеек
    for item in border_highlighting_list:
        if (item[-1] - item[0]) >= 1:
            for row in range(item[0], item[1] + 1):
                for col in range(5, 11):
                    cell = sheet_name.cell(row=row, column=col)
                    if row == item[0]:
                        if col == 5:
                            cell.border = Border(top=medium, left=medium, right=thin, bottom=thin)
                        elif col == 10:
                            cell.border = Border(top=medium, left=thin, right=medium, bottom=thin)
                        else:
                            cell.border = Border(top=medium, left=thin, right=thin, bottom=thin)
                    elif row == item[1]:
                        if col == 5:
                            cell.border = Border(top=thin, left=medium, right=thin, bottom=medium)
                        elif col == 10:
                            cell.border = Border(top=thin, left=thin, right=medium, bottom=medium)
                        else:
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=medium)
                    else:
                        if col == 5:
                            cell.border = Border(top=thin, left=medium, right=thin, bottom=thin)
                        elif col == 10:
                            cell.border = Border(top=thin, left=thin, right=medium, bottom=thin)

    sheet_name.column_dimensions.group('A', 'D', hidden=True)


def cell_style(cell, column):
    if column <= 18 or column == 22:
        cell.style = 'info'
    elif column == 19:
        cell.style = 'date'
    elif column == 20 or column == 21:
        cell.style = 'white'
    if column >= 9:
        cell.alignment = Alignment(horizontal='center', vertical='center')


def create_output_file(output_name='unnamed'):
    global output_book
    output_book = openpyxl.Workbook()
    output_book.remove(output_book.active)
    output_book.create_sheet(title='Processed_data', index=0)
    global output_sheet
    output_sheet = output_book['Processed_data']

    # Создание стиля для шапки
    ns_header = NamedStyle(name='header')
    ns_header.font = Font(name='Arial', bold=True, size=10)
    ns_header.fill = PatternFill("solid", fgColor="D6E5CB")
    thin = Side(border_style="thin", color="000000")
    ns_header.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ns_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    output_book.add_named_style(ns_header)

    # Добавление и стилизация шапки
    create_sheet_header(sheet_name=output_sheet)

    # Создание дополнительных стилей
    ns_info = NamedStyle(name='info')
    thin = Side(border_style="thin", color="000000")
    ns_info.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ns_info.font = Font(name='Arial', bold=False, size=8)
    output_book.add_named_style(ns_info)

    ns_date = NamedStyle(name='date')
    thin = Side(border_style="thin", color="000000")
    ns_date.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ns_date.font = Font(name='Arial', bold=True, size=10)
    ns_date.alignment = Alignment(horizontal='center', vertical='center')
    output_book.add_named_style(ns_date)

    ns_white = NamedStyle(name='white')
    ns_white.font = Font(name='Arial', bold=False, size=8)
    ns_white.alignment = Alignment(horizontal='center', vertical='center')
    ns_white.fill = PatternFill("solid", fgColor="FFFFFF")
    output_book.add_named_style(ns_white)

    # Копирование строк без изменений с применением стилей
    start_row = 7
    for row in range(start_row, len(row_indexes) + start_row):
        for col in range(1, len(column_indexes) + 1):
            info = input_sheet.cell(row=row_indexes[row - start_row], column=column_indexes[col - 1]).value
            if 9 <= col <= 18 and info is None:
                info = 0
            cell = output_sheet.cell(row=row, column=col)
            cell.value = info
            cell.style = 'info'
            if col >= 9:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Добавление столбца с датами (столбец №19)
    for row in range(start_row, len(row_indexes) + start_row):
        info = date_dict[row_indexes[row - start_row]]
        cell = output_sheet.cell(row=row, column=19)
        cell.value = info
        cell.style = 'date'

    # Добавление данных из ОТК (столбец №20)
    for row in range(start_row, len(row_indexes) + start_row):
        info = otk_dict.get(output_sheet.cell(row=row, column=5).value)
        if info == 0:
            info = None
        cell = output_sheet.cell(row=row, column=20)
        cell.value = info

    # Стилизация столбцом 20 и 21
    for row in range(start_row, len(row_indexes) + start_row):
        for col in range(20, 22):
            output_sheet.cell(row=row, column=col).style = 'white'

    # Добавление данных из unshipped_dict (столбец №22)
    for row in range(start_row, len(row_indexes) + start_row):
        info = unshipped_dict.get(output_sheet.cell(row=row, column=5).value)
        if info == 0:
            info = None
        cell = output_sheet.cell(row=row, column=22)
        cell.value = info
        cell.style = 'info'
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Выделение цветом номенклатур с маленьким запасом
    fill_small_stock(sheet_name=output_sheet, start_row=start_row)

    # Разнесение информации по разным листам итогового файла
    if output_name == 'abutments':
        sheet_names_list = [
            'Прямой, временный',
            'Приливаемый'
        ]

        for i, name in enumerate(sheet_names_list):
            output_book.create_sheet(title=name, index=i)
            create_sheet_header(sheet_name=output_book[name])

        r1, r2 = [start_row for _ in range(2)]

        for row in range(start_row, output_sheet.max_row + 1):
            if output_sheet.cell(row=row, column=1).value == 'Абатмент приливаемый':
                for col in range(1, 23):
                    cell = output_book['Приливаемый'].cell(row=r2, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r2 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['Прямой, временный'].cell(row=r1, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r1 += 1

        for name in sheet_names_list:
            fill_small_stock(sheet_name=output_book[name], start_row=start_row)
            separation_nomenclatures(sheet_name=output_book[name], start_row=start_row)
            create_sheet_result(sheet_name=output_book[name],
                                start_row=start_row,
                                end_row=output_book[name].max_row)
    elif output_name == 'screws':
        sheet_names_list = [
            'Винты LM',
            'Собств. разработка',
            'Zirkonzahn',
            'NT-trading',
            'SIRONA',
            'Для трансферов',
            'Лабораторные винты LM'
        ]

        for i, name in enumerate(sheet_names_list):
            output_book.create_sheet(title=name, index=i)
            create_sheet_header(sheet_name=output_book[name])

        r1, r2, r3, r4, r5, r6, r7 = [start_row for _ in range(7)]

        for row in range(start_row, output_sheet.max_row + 1):
            if 'трансфер' in output_sheet.cell(row=row, column=1).value.lower():
                for col in range(1, 23):
                    cell = output_book['Для трансферов'].cell(row=r6, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r6 += 1
            elif 'SIRONA' == output_sheet.cell(row=row, column=2).value:
                for col in range(1, 23):
                    cell = output_book['SIRONA'].cell(row=r5, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r5 += 1
            elif 'Аналог NT-Traiding' == output_sheet.cell(row=row, column=2).value:
                for col in range(1, 23):
                    cell = output_book['NT-trading'].cell(row=r4, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r4 += 1
            elif 'ZIRKONZAHN' == output_sheet.cell(row=row, column=2).value:
                for col in range(1, 23):
                    cell = output_book['Zirkonzahn'].cell(row=r3, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r3 += 1
            elif ('3D' in output_sheet.cell(row=row, column=1).value) or \
                 ('Втулка сварного винта' == output_sheet.cell(row=row, column=1).value) or \
                 ('Пин' == output_sheet.cell(row=row, column=1).value) or \
                 ('угл' in output_sheet.cell(row=row, column=1).value) or \
                 ('Винт LM (собств. разр.)' in output_sheet.cell(row=row, column=5).value):
                for col in range(1, 23):
                    cell = output_book['Собств. разработка'].cell(row=r2, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r2 += 1
            elif 'Винт LM (копия оригинала)' in output_sheet.cell(row=row, column=5).value:
                for col in range(1, 23):
                    cell = output_book['Винты LM'].cell(row=r1, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r1 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['Лабораторные винты LM'].cell(row=r7, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r7 += 1

        for name in sheet_names_list:
            fill_small_stock(sheet_name=output_book[name], start_row=start_row)
            if name == 'Лабораторные винты LM':
                output_book[name].column_dimensions.group('A', 'D', hidden=True)
                create_sheet_result(sheet_name=output_book[name],
                                    start_row=start_row,
                                    end_row=output_book[name].max_row + 1)
                continue
            separation_nomenclatures(sheet_name=output_book[name], start_row=start_row)
            create_sheet_result(sheet_name=output_book[name],
                                start_row=start_row,
                                end_row=output_book[name].max_row)
    elif output_name == 'titanium_base':
        sheet_names_list = [
            'Patch',
            'Flat с насечками (ИМ Ортос)',
            'Half (ИМ Абатменты.ру)',
            'GEO Bell',
            'GEO Step',
            'Arum',
            'Остальное'
        ]

        for i, name in enumerate(sheet_names_list):
            output_book.create_sheet(title=name, index=i)
            create_sheet_header(sheet_name=output_book[name])

        row_indexes_list = [[] for _ in range(10)]

        for row in range(start_row, output_sheet.max_row + 1):
            if 'Patch' == output_sheet.cell(row=row, column=2).value:
                if 'ТО bridge' == output_sheet.cell(row=row, column=1).value:
                    row_indexes_list[0].append(row)
                else:
                    row_indexes_list[1].append(row)
            elif 'Flat' == output_sheet.cell(row=row, column=2).value:
                if 'ТО bridge' == output_sheet.cell(row=row, column=1).value:
                    row_indexes_list[2].append(row)
                else:
                    row_indexes_list[3].append(row)
            elif 'Half' == output_sheet.cell(row=row, column=2).value:
                row_indexes_list[4].append(row)
            elif 'Bell GEO' == output_sheet.cell(row=row, column=2).value:
                if 'ТО bridge' == output_sheet.cell(row=row, column=1).value:
                    row_indexes_list[5].append(row)
                else:
                    row_indexes_list[6].append(row)
            elif 'Step GEO' == output_sheet.cell(row=row, column=2).value:
                row_indexes_list[7].append(row)
            elif 'Step ARUM' == output_sheet.cell(row=row, column=2).value:
                row_indexes_list[8].append(row)
            else:
                row_indexes_list[9].append(row)

        r1, r2, r3, r4, r5, r6, r7 = [start_row for _ in range(7)]

        p_dict = {}
        for row in row_indexes_list[0]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                        output_sheet.cell(row=row, column=8).value[:-2])] = \
                    [output_sheet.cell(row=row, column=col).value for col in range(9, 19)]

        cell = output_book['Patch'].cell(row=r1, column=5)
        cell.value = 'Мостовидные'
        for col in range(1, 23):
            cell = output_book['Patch'].cell(row=r1, column=col)
            if col == 5:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell_style(cell=cell, column=col)
        r1 += 1

        for row in row_indexes_list[0]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                continue
            elif p_dict.get((output_sheet.cell(row=row, column=5).value.split()[0][:6],
                             output_sheet.cell(row=row, column=8).value), 0) != 0:
                for col in range(1, 23):
                    cell = output_book['Patch'].cell(row=r1, column=col)
                    if 9 <= col <= 18:
                        info = output_sheet.cell(row=row, column=col).value + \
                               p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                                       output_sheet.cell(row=row, column=8).value)][col-9]
                    else:
                        info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r1 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['Patch'].cell(row=r1, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r1 += 1

        cell = output_book['Patch'].cell(row=r1, column=5)
        cell.value = 'Одиночные'
        for col in range(1, 23):
            cell = output_book['Patch'].cell(row=r1, column=col)
            if col == 5:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell_style(cell=cell, column=col)
        r1 += 1

        p_dict = {}
        for row in row_indexes_list[1]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                        output_sheet.cell(row=row, column=8).value[:-2])] = \
                    [output_sheet.cell(row=row, column=col).value for col in range(9, 19)]

        for row in row_indexes_list[1]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                continue
            elif p_dict.get((output_sheet.cell(row=row, column=5).value.split()[0][:6],
                             output_sheet.cell(row=row, column=8).value), 0) != 0:
                for col in range(1, 23):
                    cell = output_book['Patch'].cell(row=r1, column=col)
                    if 9 <= col <= 18:
                        info = output_sheet.cell(row=row, column=col).value + \
                               p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                                       output_sheet.cell(row=row, column=8).value)][col - 9]
                    else:
                        info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r1 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['Patch'].cell(row=r1, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r1 += 1

        p_dict = {}
        for row in row_indexes_list[2]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                        output_sheet.cell(row=row, column=8).value[:-2])] = \
                    [output_sheet.cell(row=row, column=col).value for col in range(9, 19)]

        cell = output_book['Flat с насечками (ИМ Ортос)'].cell(row=r2, column=5)
        cell.value = 'Мостовидные'
        for col in range(1, 23):
            cell = output_book['Flat с насечками (ИМ Ортос)'].cell(row=r2, column=col)
            if col == 5:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell_style(cell=cell, column=col)
        r2 += 1

        for row in row_indexes_list[2]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                continue
            elif p_dict.get((output_sheet.cell(row=row, column=5).value.split()[0],
                             output_sheet.cell(row=row, column=8).value), 0) != 0:
                for col in range(1, 23):
                    cell = output_book['Flat с насечками (ИМ Ортос)'].cell(row=r2, column=col)
                    if 9 <= col <= 18:
                        info = output_sheet.cell(row=row, column=col).value + \
                               p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                                       output_sheet.cell(row=row, column=8).value)][col - 9]
                    else:
                        info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r2 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['Flat с насечками (ИМ Ортос)'].cell(row=r2, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r2 += 1

        cell = output_book['Flat с насечками (ИМ Ортос)'].cell(row=r2, column=5)
        cell.value = 'Одиночные'
        for col in range(1, 23):
            cell = output_book['Flat с насечками (ИМ Ортос)'].cell(row=r2, column=col)
            if col == 5:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell_style(cell=cell, column=col)
        r2 += 1

        p_dict = {}
        for row in row_indexes_list[3]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                        output_sheet.cell(row=row, column=8).value[:-2])] = \
                    [output_sheet.cell(row=row, column=col).value for col in range(9, 19)]

        for row in row_indexes_list[3]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                continue
            elif p_dict.get((output_sheet.cell(row=row, column=5).value.split()[0],
                             output_sheet.cell(row=row, column=8).value), 0) != 0:
                for col in range(1, 23):
                    cell = output_book['Flat с насечками (ИМ Ортос)'].cell(row=r2, column=col)
                    if 9 <= col <= 18:
                        info = output_sheet.cell(row=row, column=col).value + \
                               p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                                       output_sheet.cell(row=row, column=8).value)][col - 9]
                    else:
                        info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r2 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['Flat с насечками (ИМ Ортос)'].cell(row=r2, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r2 += 1

        p_dict = {}
        for row in row_indexes_list[4]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                        output_sheet.cell(row=row, column=8).value[:-2])] = \
                    [output_sheet.cell(row=row, column=col).value for col in range(9, 19)]

        for row in row_indexes_list[4]:
            if 'P' in output_sheet.cell(row=row, column=8).value:
                continue
            elif p_dict.get((output_sheet.cell(row=row, column=5).value.split()[0],
                             output_sheet.cell(row=row, column=8).value), 0) != 0:
                for col in range(1, 23):
                    cell = output_book['Half (ИМ Абатменты.ру)'].cell(row=r3, column=col)
                    if 9 <= col <= 18:
                        info = output_sheet.cell(row=row, column=col).value + \
                               p_dict[(output_sheet.cell(row=row, column=5).value.split()[0][:6],
                                       output_sheet.cell(row=row, column=8).value)][col - 9]
                    else:
                        info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r3 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['Half (ИМ Абатменты.ру)'].cell(row=r3, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r3 += 1

        cell = output_book['GEO Bell'].cell(row=r4, column=5)
        cell.value = 'Мостовидные'
        for col in range(1, 23):
            cell = output_book['GEO Bell'].cell(row=r4, column=col)
            if col == 5:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell_style(cell=cell, column=col)
        r4 += 1

        for row in row_indexes_list[5]:
            if output_sheet.cell(row=row, column=5).value in tb_ar_dict.keys():
                for col in range(1, 23):
                    if (1 <= col <= 8) or (19 <= col <= 22):
                        cell = output_book['GEO Bell'].cell(row=r4, column=col)
                        info = output_sheet.cell(row=row, column=col).value
                        cell.value = info
                        cell_style(cell=cell, column=col)
                    else:
                        cell = output_book['GEO Bell'].cell(row=r4, column=col)
                        info = output_sheet.cell(row=row, column=col).value + \
                               tb_ar_dict[output_sheet.cell(row=row, column=5).value][col-9]
                        cell.value = info
                        cell_style(cell=cell, column=col)
                r4 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['GEO Bell'].cell(row=r4, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r4 += 1

        cell = output_book['GEO Bell'].cell(row=r4, column=5)
        cell.value = 'Одиночные'
        for col in range(1, 23):
            cell = output_book['GEO Bell'].cell(row=r4, column=col)
            if col == 5:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell_style(cell=cell, column=col)
        r4 += 1

        for row in row_indexes_list[6]:
            if output_sheet.cell(row=row, column=5).value in tb_ar_dict.keys():
                for col in range(1, 23):
                    if (1 <= col <= 8) or (19 <= col <= 22):
                        cell = output_book['GEO Bell'].cell(row=r4, column=col)
                        info = output_sheet.cell(row=row, column=col).value
                        cell.value = info
                        cell_style(cell=cell, column=col)
                    else:
                        cell = output_book['GEO Bell'].cell(row=r4, column=col)
                        info = output_sheet.cell(row=row, column=col).value + \
                               tb_ar_dict[output_sheet.cell(row=row, column=5).value][col-9]
                        cell.value = info
                        cell_style(cell=cell, column=col)
                r4 += 1
            else:
                for col in range(1, 23):
                    cell = output_book['GEO Bell'].cell(row=r4, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r4 += 1

        for row in row_indexes_list[7]:
            for col in range(1, 23):
                cell = output_book['GEO Step'].cell(row=r5, column=col)
                info = output_sheet.cell(row=row, column=col).value
                cell.value = info
                cell_style(cell=cell, column=col)
            r5 += 1

        for row in row_indexes_list[8]:
            for col in range(1, 23):
                cell = output_book['Arum'].cell(row=r6, column=col)
                info = output_sheet.cell(row=row, column=col).value
                cell.value = info
                cell_style(cell=cell, column=col)
            r6 += 1

        for row in row_indexes_list[9]:
            for col in range(1, 23):
                cell = output_book['Остальное'].cell(row=r7, column=col)
                info = output_sheet.cell(row=row, column=col).value
                cell.value = info
                cell_style(cell=cell, column=col)
            r7 += 1

        for name in sheet_names_list:
            fill_small_stock(sheet_name=output_book[name], start_row=start_row)
            separation_nomenclatures(sheet_name=output_book[name], start_row=start_row)
            create_sheet_result(sheet_name=output_book[name],
                                start_row=start_row,
                                end_row=output_book[name].max_row)

        yellow_list = [
            '38779',
            '38780',
            '38781',
            '38711',
            '38732',
            '38735',
            '38737',
            '38755',
            '38625',
            '38626',
            '38621',
            '38622',
            '38624',
            '38606',
            '38608',
            '38617',
            '38618'
        ]
        for name in ['GEO Step', 'Arum']:
            for row in range(7, output_book[name].max_row):
                cell = output_book[name].cell(row=row, column=5)
                if cell.value is not None:
                    if cell.value.split()[0][:5] in yellow_list:
                        cell.fill = PatternFill("solid", fgColor="FFFF00")
    elif output_name == 'blanks':
        sheet_names_list = [
            'ARUM_CoCr',
            'ARUM_Ti',
            'Zirkonzahn',
            'ADM_Medentika',
            'Остальное',
        ]

        for i, name in enumerate(sheet_names_list):
            output_book.create_sheet(title=name, index=i)
            create_sheet_header(sheet_name=output_book[name])

        output_book.save(
            f'output_files/!{file_folder_path_dict[output_name][-1]}_{date_stop[:5]}-{date_start[:5]}.xlsx')

        path1 = os.path.abspath('data/blanks/blanks_LM1.xlsx')
        path2 = os.path.abspath(
            f'output_files/!{file_folder_path_dict[output_name][-1]}_{date_stop[:5]}-{date_start[:5]}.xlsx')

        xl = Dispatch("Excel.Application")
        xl.Visible = True  # You can remove this line if you don't want the Excel application to be visible

        wb1 = xl.Workbooks.Open(Filename=path1)
        wb2 = xl.Workbooks.Open(Filename=path2)

        ws1 = wb1.Worksheets(1)
        ws1.Copy(Before=wb2.Worksheets(1))

        wb2.Close(SaveChanges=True)
        xl.Quit()

        output_book = openpyxl.load_workbook(
            f'output_files/!{file_folder_path_dict[output_name][-1]}_{date_stop[:5]}-{date_start[:5]}.xlsx',)
        output_sheet = output_book['Processed_data']
        create_sheet_header(sheet_name=output_book['LM1'])
        output_book['LM1'].column_dimensions.group('A', 'D', hidden=True)

        r1, r2, r3, r4, r5, r6 = [start_row for _ in range(6)]

        for row in range(start_row, output_sheet.max_row + 1):
            if 'Для холдеров ADM / MEDENTiKA' == output_sheet.cell(row=row, column=2).value:
                for col in range(1, 23):
                    cell = output_book['ADM_Medentika'].cell(row=r5, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r5 += 1
            elif 'Для холдера ZIRKONZAHN' == output_sheet.cell(row=row, column=2).value:
                for col in range(1, 23):
                    cell = output_book['Zirkonzahn'].cell(row=r4, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r4 += 1
            elif 'Для холдера ARUM' == output_sheet.cell(row=row, column=2).value and \
                 'КХ' in output_sheet.cell(row=row, column=5).value.split()[0]:
                for col in range(1, 23):
                    cell = output_book['ARUM_CoCr'].cell(row=r2, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r2 += 1
            elif 'Для холдера ARUM' == output_sheet.cell(row=row, column=2).value:
                for col in range(1, 23):
                    cell = output_book['ARUM_Ti'].cell(row=r3, column=col)
                    info = output_sheet.cell(row=row, column=col).value
                    cell.value = info
                    cell_style(cell=cell, column=col)
                r3 += 1
            else:
                for i in range(7, output_book['LM1'].max_row+1):
                    if output_book['LM1'].cell(row=i, column=5).value == output_sheet.cell(row=row, column=5).value:
                        for col in range(1, 23):
                            cell = output_book['LM1'].cell(row=i, column=col)
                            info = output_sheet.cell(row=row, column=col).value
                            cell.value = info
                        break
                else:
                    for col in range(1, 23):
                        cell = output_book['Остальное'].cell(row=r6, column=col)
                        info = output_sheet.cell(row=row, column=col).value
                        cell.value = info
                        cell_style(cell=cell, column=col)
                    r6 += 1

        sheet_names_list.append('LM1')
        for name in sheet_names_list:
            if name == 'LM1':
                fill_small_stock(sheet_name=output_book[name], start_row=start_row)
                create_sheet_result(sheet_name=output_book[name],
                                    start_row=start_row,
                                    end_row=output_book[name].max_row+1)
            else:
                fill_small_stock(sheet_name=output_book[name], start_row=start_row)
                separation_nomenclatures(sheet_name=output_book[name], start_row=start_row)
                create_sheet_result(sheet_name=output_book[name],
                                    start_row=start_row,
                                    end_row=output_book[name].max_row)
    elif output_name == 'implants':
        # Добавление столбец №23
        output_sheet.merge_cells(start_row=4, start_column=23, end_row=6, end_column=23)
        output_sheet.cell(row=4, column=23).value = 'Необходимый остаток на 4 месяца'
        output_sheet.cell(row=4, column=23).style = 'header'
        output_sheet.column_dimensions[get_column_letter(23)].width = 20
        for row in range(7, output_sheet.max_row+1):
            info = stock_dict.get(int(output_sheet.cell(row=row, column=5).value.split()[0][:5]))
            cell = output_sheet.cell(row=row, column=23)
            cell.value = info
            cell.style = 'info'
            cell.alignment = Alignment(horizontal='center', vertical='center')

        separation_nomenclatures(sheet_name=output_sheet, start_row=start_row)
        create_sheet_result(sheet_name=output_sheet,
                            start_row=start_row,
                            end_row=output_sheet.max_row)

        green_list = [
            '51022',
            '51023',
            '51024',
            '51025',
            '51026',
            '51027'
        ]
        for row in range(7, output_sheet.max_row):
            cell = output_sheet.cell(row=row, column=5)
            if cell.value is not None:
                if cell.value.split()[0][:5] in green_list:
                    cell.fill = PatternFill("solid", fgColor="CCFFCC")
    elif output_name == 'analog':
        create_sheet_result(sheet_name=output_sheet,
                            start_row=start_row,
                            end_row=output_sheet.max_row+1)
        output_sheet.column_dimensions.group('A', 'D', hidden=True)
    else:
        separation_nomenclatures(sheet_name=output_sheet, start_row=start_row)
        create_sheet_result(sheet_name=output_sheet,
                            start_row=start_row,
                            end_row=output_sheet.max_row)

    output_book.save(f'output_files/!{file_folder_path_dict[output_name][-1]}_{date_stop[:5]}-{date_start[:5]}.xlsx')
    print(f'!{file_folder_path_dict[output_name][-1]}_{date_stop[:5]}-{date_start[:5]}.xlsx is done')


def main():
    global date_start, date_stop
    # date_start, date_stop = input('Начало периода: '), input('Конец периода: ')
    date_start, date_stop = '19.03.2022', '20.04.2022'

    for key, value in file_folder_path_dict.items():
        read_input_files(main_file=f'data/{key}/{value[0]}',
                         unshipped_file=f'data/{key}/{value[1]}')
        create_output_file(output_name=key)


if __name__ == '__main__':
    main()
